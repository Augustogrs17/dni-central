"""
exportador.py
=============
Gera o arquivo 'Projetos Preliminares - YYMMDD.xlsx' com 4 abas:

  Resumo              — dashboard com KPIs + 3 gráficos posicionados estrategicamente
  Projetos            — todos os projetos (não cancelados), com formatação condicional
  Projetos Efetivados — somente projetos efetivados, ordenados por data
  Projetos Preliminares — somente status PRELIMINAR, ordenados por DIAS (mais atrasados primeiro)
"""

import datetime
import os
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.series import SeriesLabel

# ── Paleta Metalfrio ──────────────────────────────────────────────────
_AZUL    = "0B1D2F"   # cabeçalho escuro
_TEAL    = "00B4C8"   # accent primário
_VERDE   = "00A651"   # positivo
_LARANJA = "E05C2A"   # atenção
_CINZA_C = "F7F9FC"   # zebra claro
_CINZA_B = "EEF3FA"   # zebra escuro
_BRANCO  = "FFFFFF"
_VERMELHO= "C0392B"

# ── Estilos reutilizáveis ─────────────────────────────────────────────
def _side(cor="D0D7E2", estilo="thin"):
    return Side(style=estilo, color=cor)

def _borda(cor="D0D7E2"):
    s = _side(cor)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_cor):
    return PatternFill("solid", start_color=hex_cor, end_color=hex_cor)

def _font(bold=False, cor=_BRANCO, size=10, italic=False):
    return Font(bold=bold, color=cor, size=size, italic=italic,
                name="Segoe UI")

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Helpers de data ───────────────────────────────────────────────────
_EPOCH = datetime.datetime(1899, 12, 30)

def _to_date(val: Any) -> datetime.date | None:
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, (int, float)):
        n = int(val)
        if 18000 <= n <= 73000:
            try:
                return (_EPOCH + datetime.timedelta(days=n)).date()
            except Exception:
                pass
    if isinstance(val, str):
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                return datetime.datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                pass
    return None

def _fmt_date(val: Any) -> str:
    d = _to_date(val)
    return d.strftime("%d/%m/%y") if d else ""


# ── Formatação de aba de dados ────────────────────────────────────────
def _aplicar_header(ws, cabecalhos: list[str], larguras: list[float]):
    for j, (titulo, larg) in enumerate(zip(cabecalhos, larguras), 1):
        c = ws.cell(1, j, titulo)
        c.font      = _font(bold=True, cor=_BRANCO, size=9)
        c.fill      = _fill(_AZUL)
        c.alignment = _align()
        c.border    = _borda("1C3050")
        ws.column_dimensions[get_column_letter(j)].width = larg
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _aplicar_linha(ws, row: int, valores: list, cols_data: set = None,
                   col_assunto: int = None, zebra: bool = False):
    bg = _CINZA_B if zebra else _BRANCO
    for j, val in enumerate(valores, 1):
        # Formata datas como string para evitar exibição de serial
        if cols_data and j in cols_data:
            val = _fmt_date(val) or val
        c = ws.cell(row, j, val)
        c.fill   = _fill(bg)
        c.border = _borda()
        c.font   = Font(size=9, name="Segoe UI",
                        color="1A1A1A" if j != col_assunto else "0B1D2F")
        c.alignment = _align(
            h="left" if col_assunto and j == col_assunto else "center"
        )
    ws.row_dimensions[row].height = 16


def _formatar_col_dias(ws, col: int, linha_ini: int, linha_fim: int):
    if linha_fim < linha_ini:
        return
    letra = get_column_letter(col)
    rng   = f"{letra}{linha_ini}:{letra}{linha_fim}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0"],
        fill=_fill("FDECEA"),
        font=Font(color=_VERMELHO, bold=True, size=9, name="Segoe UI"),
    ))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThanOrEqual", formula=["0"],
        fill=_fill("E8F5E9"),
        font=Font(color=_VERDE, bold=True, size=9, name="Segoe UI"),
    ))


def _auto_filter(ws, n_cols: int, ultima_linha: int):
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{ultima_linha}"


# ══════════════════════════════════════════════════════════════════════
# ABA PROJETOS
# ══════════════════════════════════════════════════════════════════════
def _criar_aba_projetos(wb, projetos: list[dict]):
    ws = wb.create_sheet("Projetos")

    cabs = [
        "DEP", "DIAS", "STATUS", "LÍDER TÉCNICO", "FASE",
        "CATEGORIA", "ASSUNTO", "MODELO", "SOLICITAÇÃO", "ESTIMATIVA",
    ]
    larg = [10, 8, 14, 16, 16, 16, 48, 18, 12, 12]
    _aplicar_header(ws, cabs, larg)

    hoje = datetime.date.today()
    for i, p in enumerate(projetos, 2):
        status   = p.get("status", "")
        est_date = _to_date(p.get("estimativa"))
        sol_date = _to_date(p.get("solicitacao"))

        if status == "EFETIVADO":
            efet_date = _to_date(p.get("data_efet"))
            dias = (hoje - efet_date).days if efet_date else None
        elif status == "PRELIMINAR":
            dias = (est_date - hoje).days if est_date else None
        else:
            dias = None

        vals = [
            p.get("dep", ""),
            dias,
            status,
            p.get("lider", ""),
            p.get("fase", ""),
            p.get("categoria", ""),
            p.get("assunto", ""),
            p.get("modelo", ""),
            _fmt_date(sol_date) if sol_date else "",
            _fmt_date(est_date) if est_date else "",
        ]
        _aplicar_linha(ws, i, vals, col_assunto=7, zebra=(i % 2 == 0))

    ultima = len(projetos) + 1
    _formatar_col_dias(ws, 2, 2, ultima)
    _auto_filter(ws, len(cabs), ultima)
    return ws


# ══════════════════════════════════════════════════════════════════════
# ABA PROJETOS EFETIVADOS
# ══════════════════════════════════════════════════════════════════════
def _criar_aba_efetivados(wb, projetos: list[dict]):
    ws = wb.create_sheet("Projetos Efetivados")

    efetivados = sorted(
        [p for p in projetos if p.get("status") == "EFETIVADO"],
        key=lambda p: _to_date(p.get("data_efet")) or datetime.date.min,
        reverse=True,
    )

    cabs = [
        "DEP", "DATA EFETIVAÇÃO", "DIAS (ATRASO)",
        "ASSUNTO", "LÍDER TÉCNICO", "CATEGORIA", "SOLICITAÇÃO", "ESTIMATIVA",
    ]
    larg = [10, 14, 13, 48, 16, 16, 12, 12]
    _aplicar_header(ws, cabs, larg)

    hoje = datetime.date.today()
    for i, p in enumerate(efetivados, 2):
        efet = _to_date(p.get("data_efet"))
        est  = _to_date(p.get("estimativa"))
        dias_atraso = (est - efet).days if efet and est else None

        vals = [
            p.get("dep", ""),
            _fmt_date(efet),
            dias_atraso,
            p.get("assunto", ""),
            p.get("lider", ""),
            p.get("categoria", ""),
            _fmt_date(_to_date(p.get("solicitacao"))),
            _fmt_date(est),
        ]
        _aplicar_linha(ws, i, vals, col_assunto=4, zebra=(i % 2 == 0))

    ultima = len(efetivados) + 1
    _formatar_col_dias(ws, 3, 2, ultima)
    _auto_filter(ws, len(cabs), ultima)
    return ws


# ══════════════════════════════════════════════════════════════════════
# ABA PROJETOS PRELIMINARES
# ══════════════════════════════════════════════════════════════════════
def _criar_aba_preliminares(wb, projetos: list[dict]):
    ws = wb.create_sheet("Projetos Preliminares")

    hoje = datetime.date.today()
    preli = []
    for p in projetos:
        if p.get("status") != "PRELIMINAR":
            continue
        est = _to_date(p.get("estimativa"))
        dias = (est - hoje).days if est else None
        preli.append({**p, "_dias": dias})

    # Mais atrasados primeiro
    preli.sort(key=lambda p: p["_dias"] if p["_dias"] is not None else 9999)

    cabs = [
        "DEP", "DIAS RESTANTES", "ESTIMATIVA", "SOLICITAÇÃO",
        "ASSUNTO", "LÍDER TÉCNICO", "FASE", "CATEGORIA",
    ]
    larg = [10, 14, 12, 12, 48, 16, 16, 16]
    _aplicar_header(ws, cabs, larg)

    for i, p in enumerate(preli, 2):
        vals = [
            p.get("dep", ""),
            p["_dias"],
            _fmt_date(_to_date(p.get("estimativa"))),
            _fmt_date(_to_date(p.get("solicitacao"))),
            p.get("assunto", ""),
            p.get("lider", ""),
            p.get("fase", ""),
            p.get("categoria", ""),
        ]
        _aplicar_linha(ws, i, vals, col_assunto=5, zebra=(i % 2 == 0))

    ultima = len(preli) + 1
    _formatar_col_dias(ws, 2, 2, ultima)
    _auto_filter(ws, len(cabs), ultima)
    return ws


# ══════════════════════════════════════════════════════════════════════
# ABA RESUMO — dashboard profissional
# ══════════════════════════════════════════════════════════════════════
def _criar_aba_resumo(wb, projetos: list[dict], resumo: dict):
    ws = wb.create_sheet("Resumo", 0)
    ws.sheet_view.showGridLines = False

    # ── Dimensões da grade ────────────────────────────────────────────
    # Cols A-B: margem | C-H: KPIs | I: sep | J-R: gráfico 1 | S: sep | T-AB: gráfico 2
    ws.column_dimensions["A"].width = 1.5
    ws.column_dimensions["B"].width = 1.5
    for c in "CDEFGH":
        ws.column_dimensions[c].width = 13
    ws.column_dimensions["I"].width = 1.5
    for c in ["J","K","L","M","N","O","P","Q","R"]:
        ws.column_dimensions[c].width = 7
    ws.column_dimensions["S"].width = 1.5
    for c in ["T","U","V","W","X","Y","Z"]:
        ws.column_dimensions[c].width = 7
    ws.column_dimensions["AA"].width = 7
    ws.column_dimensions["AB"].width = 7
    # Tabelas auxiliares
    ws.column_dimensions["AD"].width = 20
    ws.column_dimensions["AE"].width = 10

    # Altura das linhas
    ws.row_dimensions[1].height  = 10   # margem topo
    ws.row_dimensions[2].height  = 28   # título
    ws.row_dimensions[3].height  = 8    # sep
    for r in range(4, 9):               # KPI top
        ws.row_dimensions[r].height = 26
    ws.row_dimensions[9].height  = 8
    for r in range(10, 15):             # KPI bottom
        ws.row_dimensions[r].height = 26
    ws.row_dimensions[15].height = 8
    ws.row_dimensions[16].height = 8
    for r in range(17, 38):
        ws.row_dimensions[r].height = 14

    # ── Título ────────────────────────────────────────────────────────
    ws.merge_cells("C2:H2")
    c = ws["C2"]
    c.value     = "❄  ENGENHARIA DE PRODUTO — PROJETOS DNI"
    c.font      = Font(bold=True, size=14, color=_AZUL, name="Segoe UI")
    c.alignment = _align(h="left", v="center")

    hoje_str = datetime.date.today().strftime("%d/%m/%Y")
    ws.merge_cells("J2:R2")
    c2 = ws["J2"]
    c2.value     = f"Gerado em {hoje_str}  ·  Metalfrio Solutions S.A."
    c2.font      = Font(size=9, color="888888", italic=True, name="Segoe UI")
    c2.alignment = _align(h="right", v="center")

    ef = resumo.get("efetivados", {})
    ab = resumo.get("em_aberto", {})
    idx = resumo.get("indice_eng", 0)
    idx_cls = resumo.get("indice_eng_cls", ("—", "888888"))

    # ── Bloco KPI helper ─────────────────────────────────────────────
    def _kpi_bloco(row_hdr, titulo, cor_hdr, kpis):
        """
        kpis = [(label, valor, formato, cor_val), ...]
        """
        # Header do bloco
        ws.merge_cells(f"C{row_hdr}:H{row_hdr}")
        ch = ws[f"C{row_hdr}"]
        ch.value     = f"  {titulo}"
        ch.font      = Font(bold=True, size=9, color=_BRANCO, name="Segoe UI")
        ch.fill      = _fill(cor_hdr)
        ch.alignment = _align(h="left")
        ws.row_dimensions[row_hdr].height = 18

        for i, (label, valor, fmt, cor_v) in enumerate(kpis):
            r = row_hdr + 1 + i
            # Rótulo (cols C-E)
            ws.merge_cells(f"C{r}:E{r}")
            cl = ws[f"C{r}"]
            cl.value     = label
            cl.font      = Font(size=9, color="444444", name="Segoe UI")
            cl.fill      = _fill(_CINZA_C)
            cl.alignment = _align(h="left")
            cl.border    = _borda()
            # Valor (cols F-H)
            ws.merge_cells(f"F{r}:H{r}")
            cv = ws[f"F{r}"]
            cv.value          = valor
            cv.number_format  = fmt
            cv.font           = Font(bold=True, size=11, color=cor_v, name="Segoe UI")
            cv.fill           = _fill(_BRANCO)
            cv.alignment      = _align(h="right")
            cv.border         = _borda()
            ws.row_dimensions[r].height = 22

    # Bloco EFETIVADOS (linhas 4-8)
    _kpi_bloco(4, "EFETIVADOS", _AZUL, [
        ("Total Efetivados",    ef.get("total", 0),             "0",     _TEAL),
        ("No Prazo",            ef.get("pct_no_prazo", 0),      "0.0%",  _VERDE),
        ("Atrasados",           ef.get("pct_atrasados", 0),     "0.0%",  _LARANJA),
        ("Atraso Médio (dias)", ef.get("atraso_medio_dias", 0), "0.0",   _LARANJA),
    ])

    # Bloco EM ABERTO (linhas 10-14)
    _kpi_bloco(10, "EM ABERTO (PRELIMINAR)", _AZUL, [
        ("Total em Aberto",     ab.get("total", 0),             "0",     _TEAL),
        ("No Prazo",            ab.get("pct_no_prazo", 0),      "0.0%",  _VERDE),
        ("Atrasados",           ab.get("pct_atrasados", 0),     "0.0%",  _LARANJA),
        ("Atraso Médio (dias)", ab.get("atraso_medio_dias", 0), "0.0",   _LARANJA),
    ])

    # Índice ENG (linha 16)
    ws.merge_cells("C16:E16")
    ci = ws["C16"]
    ci.value     = "Índice ENG"
    ci.font      = Font(size=9, color="444444", name="Segoe UI")
    ci.fill      = _fill(_CINZA_C)
    ci.alignment = _align(h="left")
    ci.border    = _borda()
    ws.merge_cells("F16:H16")
    cv = ws["F16"]
    cv.value          = idx
    cv.number_format  = '0"/100"'
    cor_idx = _VERDE if idx >= 80 else (_LARANJA if idx >= 60 else _VERMELHO)
    cv.font           = Font(bold=True, size=12, color=cor_idx, name="Segoe UI")
    cv.fill           = _fill(_BRANCO)
    cv.alignment      = _align(h="right")
    cv.border         = _borda()
    ws.row_dimensions[16].height = 22

    # ── Tabelas auxiliares para gráficos (col AD-AE, ocultas visualmente) ──
    # Tabela 1: Status
    contagens = dict(resumo.get("contagens", []))
    status_order = ["PRELIMINAR", "CONCEITO", "CONCEITO/PRÉ ALPHA",
                    "PRÉ ALPHA", "CANCELADO", "EFETIVADOS"]
    t1_inicio = 2
    ws["AD1"] = "STATUS"; ws["AE1"] = "QTD"
    for hdr in ("AD1", "AE1"):
        ws[hdr].font = Font(bold=True, size=9, color=_BRANCO, name="Segoe UI")
        ws[hdr].fill = _fill(_AZUL)
        ws[hdr].alignment = _align()
    for i, s in enumerate(status_order, t1_inicio):
        ws.cell(i, 30, s)
        ws.cell(i, 31, contagens.get(s, 0))
    t1_fim = t1_inicio + len(status_order) - 1

    # Tabela 2: Prazo (efetivados)
    t2_inicio = t1_fim + 2
    ws.cell(t2_inicio - 1, 30, "PRAZO"); ws.cell(t2_inicio - 1, 31, "%")
    ws.cell(t2_inicio,     30, "No Prazo")
    ws.cell(t2_inicio,     31, ef.get("pct_no_prazo", 0))
    ws.cell(t2_inicio + 1, 30, "Atrasados")
    ws.cell(t2_inicio + 1, 31, ef.get("pct_atrasados", 0))
    t2_fim = t2_inicio + 1

    # Tabela 3: Categoria atrasados
    t3_inicio = t2_fim + 2
    ws.cell(t3_inicio - 1, 30, "CATEGORIA"); ws.cell(t3_inicio - 1, 31, "ATRASADOS")
    cat_list = resumo.get("atrasados_categoria", [])
    for i, (cat, qtd) in enumerate(cat_list[:8]):
        ws.cell(t3_inicio + i, 30, cat)
        ws.cell(t3_inicio + i, 31, qtd)
    t3_fim = t3_inicio + len(cat_list[:8]) - 1

    # ── GRÁFICO 1: Pizza — Distribuição de Status ──────────────────
    # Posição: J4:R16 (topo direito, ao lado dos KPIs)
    pie = PieChart()
    pie.title  = "Distribuição de Status"
    pie.style  = 10
    pie.height = 10
    pie.width  = 14
    pie.add_data(
        Reference(ws, min_col=31, min_row=t1_inicio, max_row=t1_fim),
        titles_from_data=False,
    )
    pie.set_categories(
        Reference(ws, min_col=30, min_row=t1_inicio, max_row=t1_fim)
    )
    # Cores por fatia
    cores_pie = ["00B4C8","0B1D2F","6B7280","A78BFA","E05C2A","00A651"]
    for i, cor in enumerate(cores_pie[:len(status_order)]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = cor
        pie.series[0].data_points.append(pt)
    dl = DataLabelList()
    dl.showPercent  = True
    dl.showCatName  = True
    dl.showVal      = False
    dl.showSerName  = False
    dl.showLegendKey= False
    dl.numFmt       = "0%"
    pie.dataLabels  = dl
    pie.legend      = None
    ws.add_chart(pie, "J4")

    # ── GRÁFICO 2: Donut — No Prazo × Atrasados ───────────────────
    # Posição: T4:AB16 (topo extremo direito)
    donut = PieChart()
    donut.title  = "Efetivados: Prazo"
    donut.style  = 10
    donut.height = 10
    donut.width  = 12
    try:
        donut.type = "doughnut"
    except Exception:
        pass
    donut.add_data(
        Reference(ws, min_col=31, min_row=t2_inicio, max_row=t2_fim),
        titles_from_data=False,
    )
    donut.set_categories(
        Reference(ws, min_col=30, min_row=t2_inicio, max_row=t2_fim)
    )
    for i, cor in enumerate([_VERDE, _LARANJA]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = cor
        donut.series[0].data_points.append(pt)
    dl2 = DataLabelList()
    dl2.showPercent   = True
    dl2.showVal       = False
    dl2.showSerName   = False
    dl2.showLegendKey = False
    dl2.numFmt        = "0%"
    donut.dataLabels  = dl2
    donut.legend.position = "b"
    ws.add_chart(donut, "T4")

    # ── GRÁFICO 3: Barras horizontais — Atrasados por Categoria ───
    # Posição: J17:AB36 (faixa inferior, largura completa)
    bar = BarChart()
    bar.type   = "bar"
    bar.title  = "Projetos Atrasados por Categoria"
    bar.style  = 10
    bar.height = 11
    bar.width  = 28
    bar.add_data(
        Reference(ws, min_col=31, min_row=t3_inicio, max_row=t3_fim),
        titles_from_data=False,
    )
    bar.set_categories(
        Reference(ws, min_col=30, min_row=t3_inicio, max_row=t3_fim)
    )
    bar.series[0].graphicalProperties.solidFill = _TEAL
    bar.series[0].graphicalProperties.ln.solidFill = _TEAL
    dl3 = DataLabelList()
    dl3.showVal = True
    bar.dataLabels  = dl3
    bar.legend      = None
    bar.y_axis.majorGridlines = None
    ws.add_chart(bar, "J17")

    # ── Tabela de Líderes (visível) — col C-H, abaixo dos KPIs ───
    r_lid = 18
    ws.merge_cells(f"C{r_lid}:H{r_lid}")
    ch = ws[f"C{r_lid}"]
    ch.value     = "  PROJETOS ATIVOS POR LÍDER TÉCNICO"
    ch.font      = Font(bold=True, size=9, color=_BRANCO, name="Segoe UI")
    ch.fill      = _fill(_AZUL)
    ch.alignment = _align(h="left")
    ws.row_dimensions[r_lid].height = 18

    lideres = resumo.get("projetos_por_lider", [])
    for i, (login, qtd) in enumerate(lideres[:12]):
        r = r_lid + 1 + i
        ws.merge_cells(f"C{r}:F{r}")
        cl = ws.cell(r, 3, login)
        cl.font      = Font(size=9, color="333333", name="Segoe UI")
        cl.fill      = _fill(_CINZA_C if i % 2 == 0 else _BRANCO)
        cl.alignment = _align(h="left")
        cl.border    = _borda()
        ws.merge_cells(f"G{r}:H{r}")
        cv = ws.cell(r, 7, qtd)
        cv.font      = Font(bold=True, size=9, color=_TEAL, name="Segoe UI")
        cv.fill      = _fill(_CINZA_C if i % 2 == 0 else _BRANCO)
        cv.alignment = _align(h="center")
        cv.border    = _borda()
        ws.row_dimensions[r].height = 16

    return ws


# ══════════════════════════════════════════════════════════════════════
# FUNÇÃO PRINCIPAL
# ══════════════════════════════════════════════════════════════════════
def gerar_excel(projetos: list[dict], resumo: dict, pasta_destino: str) -> str:
    """
    Gera 'Projetos Preliminares - YYMMDD.xlsx' em pasta_destino.
    Retorna o caminho do arquivo gerado.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove aba padrão

    _criar_aba_resumo(wb, projetos, resumo)
    _criar_aba_projetos(wb, projetos)
    _criar_aba_efetivados(wb, projetos)
    _criar_aba_preliminares(wb, projetos)

    hoje     = datetime.date.today().strftime("%y%m%d")
    nome     = f"Projetos Preliminares - {hoje}.xlsx"
    caminho  = os.path.join(pasta_destino, nome)
    wb.save(caminho)
    return caminho


# ══════════════════════════════════════════════════════════════════════
# CÁLCULO DO RESUMO (independente do Excel)
# ══════════════════════════════════════════════════════════════════════
def calcular_resumo(projetos: list[dict]) -> dict:
    hoje = datetime.date.today()

    efet_total = efet_prazo = 0
    efet_atrasos = []
    ab_total = ab_prazo = 0
    ab_atrasos = []
    contagens   = {}
    cat_atras   = {}
    proj_lider  = {}
    top_atras   = []
    prox_estim  = []

    STATUS_TEXTUAIS = ["PRELIMINAR","CONCEITO","CONCEITO/PRÉ ALPHA",
                       "PRÉ ALPHA","CANCELADO"]

    for p in projetos:
        status   = p.get("status", "")
        lider    = p.get("lider", "(sem líder)")
        cat      = p.get("categoria", "(sem categoria)")
        dep      = p.get("dep", "")
        est_date = _to_date(p.get("estimativa"))

        if status == "EFETIVADO":
            efet_total += 1
            efet_date  = _to_date(p.get("data_efet"))
            if efet_date and est_date:
                dias = (est_date - efet_date).days
                if dias >= 0:
                    efet_prazo += 1
                else:
                    efet_atrasos.append(-dias)
                    cat_atras[cat] = cat_atras.get(cat, 0) + 1
        else:
            contagens[status] = contagens.get(status, 0) + 1
            if status not in STATUS_TEXTUAIS:
                pass
            if lider:
                proj_lider[lider] = proj_lider.get(lider, 0) + 1

            if status == "PRELIMINAR":
                ab_total += 1
                if est_date:
                    dias = (est_date - hoje).days
                    if dias >= 0:
                        ab_prazo += 1
                        prox_estim.append((dep, dias))
                    else:
                        ab_atrasos.append(-dias)
                        cat_atras[cat] = cat_atras.get(cat, 0) + 1
                        top_atras.append((dep, -dias))

    def _bloco(total, prazo, atrasos):
        atrasados = total - prazo
        return {
            "total":            total,
            "no_prazo":         prazo,
            "atrasados":        atrasados,
            "pct_no_prazo":     prazo / total if total else 0,
            "pct_atrasados":    atrasados / total if total else 0,
            "atraso_medio_dias": round(sum(atrasos)/len(atrasos), 1) if atrasos else 0,
        }

    bloco_ef = _bloco(efet_total, efet_prazo, efet_atrasos)
    bloco_ab = _bloco(ab_total,   ab_prazo,   ab_atrasos)

    # Índice ENG
    atr_norm     = max(0.0, 1.0 - bloco_ab["atraso_medio_dias"] / 90.0)
    n_crit       = sum(1 for _, d in prox_estim if d <= 7)
    crit_score   = max(0.0, 1.0 - (n_crit / max(ab_total, 1)) / 0.10)
    indice       = round(
        bloco_ef["pct_no_prazo"] * 30 +
        bloco_ab["pct_no_prazo"] * 40 +
        atr_norm * 20 +
        crit_score * 10
    )
    indice = max(0, min(100, indice))
    if indice >= 80:   cls = ("Saudável", _VERDE)
    elif indice >= 60: cls = ("Atenção",  _LARANJA)
    else:              cls = ("Crítico",  _VERMELHO)

    contagens["EFETIVADOS"] = efet_total
    contagens["TOTAL"]      = len(projetos)

    return {
        "efetivados":           bloco_ef,
        "em_aberto":            bloco_ab,
        "contagens":            list(contagens.items()),
        "atrasados_categoria":  sorted(cat_atras.items(), key=lambda x: -x[1]),
        "projetos_por_lider":   sorted(proj_lider.items(), key=lambda x: -x[1]),
        "top_atrasados":        sorted(top_atras, key=lambda x: -x[1])[:5],
        "proximas_estimativas": sorted(prox_estim, key=lambda x: x[1])[:5],
        "indice_eng":           indice,
        "indice_eng_cls":       cls,
    }
