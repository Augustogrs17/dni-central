"""
extrator.py
===========
Extração de metadados dos arquivos .xlsm de projeto.

Campos extraídos (aba 'Controle de Projeto'):
  Q15 → dep
  B11 → assunto
  Q17 → status   (str ou data serial quando EFETIVADO)
  B17 → lider
  B15 → modelo
  D17 → fase
  D15 → categoria
"""

import os
import zipfile
import re
import datetime
from pathlib import Path
from typing import Any

EXTENSOES_VALIDAS = {".xlsx", ".xlsm"}

CAMPOS_PROJETO = {
    "dep":       ("Controle de Projeto", "Q15"),
    "assunto":   ("Controle de Projeto", "B11"),
    "status_raw":("Controle de Projeto", "Q17"),
    "lider":     ("Controle de Projeto", "B17"),
    "modelo":    ("Controle de Projeto", "B15"),
    "fase":      ("Controle de Projeto", "D17"),
    "categoria": ("Controle de Projeto", "D15"),
}

_RE_SHEET_TAG     = re.compile(r'<sheet\b[^>]*name="([^"]*)"[^>]*r:id="([^"]*)"', re.IGNORECASE)
_RE_SHEET_TAG_ALT = re.compile(r'<sheet\b[^>]*r:id="([^"]*)"[^>]*name="([^"]*)"', re.IGNORECASE)
_RE_RELATIONSHIP  = re.compile(r'<Relationship\b[^>]*/?>',)
_RE_ATTR          = re.compile(r'(\w+)="([^"]*)"')
_RE_SHARED_SI     = re.compile(r'<si\b[^>]*>(.*?)</si>', re.DOTALL)
_RE_T_TAG         = re.compile(r'<t\b[^>]*>(.*?)</t>', re.DOTALL)


def _unescape(s: str) -> str:
    return (s.replace("&amp;", "&").replace("&lt;", "<")
             .replace("&gt;", ">").replace("&quot;", '"')
             .replace("&apos;", "'"))


def _parse_shared_strings(xml_bytes: bytes) -> list:
    text = xml_bytes.decode("utf-8", errors="replace")
    out = []
    for m in _RE_SHARED_SI.finditer(text):
        parts = _RE_T_TAG.findall(m.group(1))
        out.append(_unescape("".join(parts)))
    return out


def _build_sheet_map(wb_xml: bytes, rels_xml: bytes) -> dict:
    wb_text = wb_xml.decode("utf-8", errors="replace")
    name_to_rid = {}
    for m in _RE_SHEET_TAG.finditer(wb_text):
        name_to_rid[_unescape(m.group(1))] = m.group(2)
    if not name_to_rid:
        for m in _RE_SHEET_TAG_ALT.finditer(wb_text):
            name_to_rid[_unescape(m.group(2))] = m.group(1)

    rels_text = rels_xml.decode("utf-8", errors="replace")
    rid_to_target = {}
    for tag in _RE_RELATIONSHIP.findall(rels_text):
        attrs = dict(_RE_ATTR.findall(tag))
        if "Id" in attrs and "Target" in attrs:
            rid_to_target[attrs["Id"]] = attrs["Target"]

    result = {}
    for name, rid in name_to_rid.items():
        target = rid_to_target.get(rid, "")
        if not target:
            continue
        if target.startswith("/"):
            path = target.lstrip("/")
        elif target.startswith("xl/"):
            path = target
        else:
            path = "xl/" + target
        result[name] = path
    return result


def _extract_cell(sheet_xml: bytes, cell_ref: str, shared: list) -> Any:
    text = sheet_xml.decode("utf-8", errors="replace")
    pat = re.compile(
        r'<c\s+[^>]*\br="' + re.escape(cell_ref) + r'"[^>]*?(?:/>|>(.*?)</c>)',
        re.DOTALL,
    )
    m = pat.search(text)
    if not m:
        return None

    tag_m = re.search(
        r'<c\s+[^>]*\br="' + re.escape(cell_ref) + r'"[^>]*?(?=/>|>)', text
    )
    full_tag = tag_m.group(0) if tag_m else ""
    t_attr = re.search(r'\bt="([^"]*)"', full_tag)
    cell_type = t_attr.group(1) if t_attr else None
    inner = m.group(1) or ""

    v_m = re.search(r'<v>(.*?)</v>', inner, re.DOTALL)
    raw = v_m.group(1) if v_m else None
    if raw is not None and not raw.strip():
        raw = None

    if raw is None:
        is_m = re.search(r'<is>(.*?)</is>', inner, re.DOTALL)
        if is_m:
            return _unescape("".join(_RE_T_TAG.findall(is_m.group(1))))
        return None

    if cell_type == "s":
        try:
            return shared[int(raw)]
        except Exception:
            return None
    if cell_type in ("str", "inlineStr"):
        return _unescape(raw)
    if cell_type == "b":
        return raw == "1"

    try:
        num = float(raw)
    except ValueError:
        return _unescape(raw)

    return int(num) if num == int(num) else num


# ── Status resolution ─────────────────────────────────────────────────

_EXCEL_EPOCH = datetime.datetime(1899, 12, 30)

def _resolver_status(raw: Any) -> tuple:
    """
    Retorna (status_str, data_efetivacao_str | None).
    Q17 pode ser:
      - str  → "PRELIMINAR", "CANCELADO", "STAND BY", "CONCEITO", etc.
      - int/float (serial Excel) → projeto efetivado; converte para data
    """
    if raw is None:
        return ("INDEFINIDO", None)

    if isinstance(raw, str):
        s = raw.strip().upper()
        return (s, None)

    # Numérico → data de efetivação
    if isinstance(raw, (int, float)):
        n = int(raw)
        if 18000 <= n <= 73000:
            try:
                data = (_EXCEL_EPOCH + datetime.timedelta(days=n)).date()
                return ("EFETIVADO", data.strftime("%d/%m/%Y"))
            except Exception:
                pass
        return ("EFETIVADO", str(raw))

    if isinstance(raw, (datetime.date, datetime.datetime)):
        if isinstance(raw, datetime.datetime):
            raw = raw.date()
        return ("EFETIVADO", raw.strftime("%d/%m/%Y"))

    return (str(raw).strip().upper(), None)


def _ler_projeto_xml(caminho: str) -> dict | None:
    try:
        with zipfile.ZipFile(caminho, "r") as z:
            names = set(z.namelist())
            if "xl/workbook.xml" not in names:
                return None

            ss = (_parse_shared_strings(z.read("xl/sharedStrings.xml"))
                  if "xl/sharedStrings.xml" in names else [])
            sheet_map = _build_sheet_map(
                z.read("xl/workbook.xml"),
                z.read("xl/_rels/workbook.xml.rels"),
            )

            aba_nome = "Controle de Projeto"
            sheet_path = sheet_map.get(aba_nome)
            if not sheet_path or sheet_path not in names:
                return None

            ws_bytes = z.read(sheet_path)
            resultado = {}
            for campo, (_, celula) in CAMPOS_PROJETO.items():
                resultado[campo] = _extract_cell(ws_bytes, celula, ss)

            return resultado
    except Exception:
        return None


def _ler_projeto_openpyxl(caminho: str) -> dict | None:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
        try:
            aba_nome = "Controle de Projeto"
            if aba_nome not in wb.sheetnames:
                return None
            ws = wb[aba_nome]
            resultado = {}
            for campo, (_, celula) in CAMPOS_PROJETO.items():
                try:
                    resultado[campo] = ws[celula].value
                except Exception:
                    resultado[campo] = None
            return resultado
        finally:
            wb.close()
    except Exception:
        return None


def listar_projetos(pasta: str) -> list:
    if not os.path.isdir(pasta):
        raise FileNotFoundError(f"Pasta não encontrada: {pasta}")

    arquivos = [
        entry.path
        for entry in os.scandir(pasta)
        if entry.is_file()
        and Path(entry.name).suffix.lower() in EXTENSOES_VALIDAS
        and not entry.name.startswith("~$")
    ]

    projetos = []
    for caminho in arquivos:
        dados = _ler_projeto_xml(caminho) or _ler_projeto_openpyxl(caminho)
        if not dados:
            continue

        status, data_efet = _resolver_status(dados.get("status_raw"))

        # Ignora cancelados
        if status == "CANCELADO":
            continue

        proj = {
            "dep":          str(dados.get("dep") or "").strip(),
            "assunto":      str(dados.get("assunto") or "").strip(),
            "status":       status,
            "data_efet":    data_efet,          # só para EFETIVADO
            "lider":        str(dados.get("lider") or "").strip(),
            "modelo":       str(dados.get("modelo") or "").strip(),
            "fase":         str(dados.get("fase") or "").strip(),
            "categoria":    str(dados.get("categoria") or "").strip(),
            "arquivo":      os.path.basename(caminho),
            "caminho":      caminho,
        }
        projetos.append(proj)

    projetos.sort(key=lambda p: str(p.get("dep") or ""))
    return projetos
