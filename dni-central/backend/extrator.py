"""
extrator.py
===========
Extração de metadados dos arquivos .xlsm de projeto.

Step 1 — listar_projetos():
  Varre a pasta, lê células fixas de cada projeto via XML rápido
  (fallback openpyxl) e retorna lista de dicts com metadados.

Campos extraídos (aba 'Controle de Projeto'):
  Q15 → dep          (ex: "0002/26")
  B11 → assunto      (nome do projeto)
  Q17 → status       (ex: "PRELIMINAR", "CANCELADO")
  B17 → lider        (líder técnico)
  B15 → modelo
  D17 → fase         (ex: "ALPHA 0%", "BETA+PILOTO")
  D15 → categoria
"""

import os
import zipfile
import re
from pathlib import Path
from typing import Any

EXTENSOES_VALIDAS = {".xlsx", ".xlsm"}

# ── Mapeamento de células ─────────────────────────────────────────────
CAMPOS_PROJETO = {
    "dep":       ("Controle de Projeto", "Q15"),
    "assunto":   ("Controle de Projeto", "B11"),
    "status":    ("Controle de Projeto", "Q17"),
    "lider":     ("Controle de Projeto", "B17"),
    "modelo":    ("Controle de Projeto", "B15"),
    "fase":      ("Controle de Projeto", "D17"),
    "categoria": ("Controle de Projeto", "D15"),
}

# ── Regex pré-compilados ──────────────────────────────────────────────
_RE_SHEET_TAG     = re.compile(r'<sheet\b[^>]*name="([^"]*)"[^>]*r:id="([^"]*)"', re.IGNORECASE)
_RE_SHEET_TAG_ALT = re.compile(r'<sheet\b[^>]*r:id="([^"]*)"[^>]*name="([^"]*)"', re.IGNORECASE)
_RE_RELATIONSHIP  = re.compile(r'<Relationship\b[^>]*/?>',)
_RE_ATTR          = re.compile(r'(\w+)="([^"]*)"')
_RE_SHARED_SI     = re.compile(r'<si\b[^>]*>(.*?)</si>', re.DOTALL)
_RE_T_TAG         = re.compile(r'<t\b[^>]*>(.*?)</t>', re.DOTALL)
_RE_CELL_REF      = re.compile(r'([A-Z]+)(\d+)')


def _unescape(s: str) -> str:
    return (s.replace("&amp;", "&").replace("&lt;", "<")
             .replace("&gt;", ">").replace("&quot;", '"')
             .replace("&apos;", "'"))


def _parse_shared_strings(xml_bytes: bytes) -> list[str]:
    text = xml_bytes.decode("utf-8", errors="replace")
    out = []
    for m in _RE_SHARED_SI.finditer(text):
        parts = _RE_T_TAG.findall(m.group(1))
        out.append(_unescape("".join(parts)))
    return out


def _build_sheet_map(wb_xml: bytes, rels_xml: bytes) -> dict[str, str]:
    """name → zip path (ex: 'xl/worksheets/sheet2.xml')"""
    wb_text = wb_xml.decode("utf-8", errors="replace")
    name_to_rid: dict[str, str] = {}
    for m in _RE_SHEET_TAG.finditer(wb_text):
        name_to_rid[_unescape(m.group(1))] = m.group(2)
    if not name_to_rid:
        for m in _RE_SHEET_TAG_ALT.finditer(wb_text):
            name_to_rid[_unescape(m.group(2))] = m.group(1)

    rels_text = rels_xml.decode("utf-8", errors="replace")
    rid_to_target: dict[str, str] = {}
    for tag in _RE_RELATIONSHIP.findall(rels_text):
        attrs = dict(_RE_ATTR.findall(tag))
        if "Id" in attrs and "Target" in attrs:
            rid_to_target[attrs["Id"]] = attrs["Target"]

    result: dict[str, str] = {}
    for name, rid in name_to_rid.items():
        target = rid_to_target.get(rid, "")
        if not target:
            continue
        if target.startswith("/"):
            path = target.lstrip("/")
        elif target.startswith("xl/"):
            path = target
        else:
            path = "xl/" + target
        result[name] = path
    return result


def _extract_cell(sheet_xml: bytes, cell_ref: str, shared: list[str]) -> Any:
    """Extrai valor de uma célula via regex (sem DOM completo)."""
    text = sheet_xml.decode("utf-8", errors="replace")
    pat = re.compile(
        r'<c\s+[^>]*\br="' + re.escape(cell_ref) + r'"[^>]*?(?:/>|>(.*?)</c>)',
        re.DOTALL,
    )
    m = pat.search(text)
    if not m:
        return None

    tag_m = re.search(
        r'<c\s+[^>]*\br="' + re.escape(cell_ref) + r'"[^>]*?(?=/>|>)', text
    )
    full_tag = tag_m.group(0) if tag_m else ""
    t_attr = re.search(r'\bt="([^"]*)"', full_tag)
    cell_type = t_attr.group(1) if t_attr else None
    inner = m.group(1) or ""

    v_m = re.search(r'<v>(.*?)</v>', inner, re.DOTALL)
    raw = v_m.group(1) if v_m else None
    if raw is not None and not raw.strip():
        raw = None

    if raw is None:
        is_m = re.search(r'<is>(.*?)</is>', inner, re.DOTALL)
        if is_m:
            return _unescape("".join(_RE_T_TAG.findall(is_m.group(1))))
        return None

    if cell_type == "s":
        try:
            return shared[int(raw)]
        except Exception:
            return None
    if cell_type in ("str", "inlineStr"):
        return _unescape(raw)
    if cell_type == "b":
        return raw == "1"

    try:
        num = float(raw)
    except ValueError:
        return _unescape(raw)

    return int(num) if num == int(num) else num


def _ler_projeto_xml(caminho: str) -> dict | None:
    """Fast path: lê células via ZIP/XML sem montar o modelo openpyxl."""
    try:
        with zipfile.ZipFile(caminho, "r") as z:
            names = set(z.namelist())
            if "xl/workbook.xml" not in names:
                return None

            ss = (_parse_shared_strings(z.read("xl/sharedStrings.xml"))
                  if "xl/sharedStrings.xml" in names else [])
            sheet_map = _build_sheet_map(
                z.read("xl/workbook.xml"),
                z.read("xl/_rels/workbook.xml.rels"),
            )

            aba_nome = "Controle de Projeto"
            sheet_path = sheet_map.get(aba_nome)
            if not sheet_path or sheet_path not in names:
                return None

            ws_bytes = z.read(sheet_path)
            resultado: dict[str, Any] = {}
            for campo, (_, celula) in CAMPOS_PROJETO.items():
                resultado[campo] = _extract_cell(ws_bytes, celula, ss)

            return resultado
    except Exception:
        return None


def _ler_projeto_openpyxl(caminho: str) -> dict | None:
    """Fallback via openpyxl (mais lento, mas robusto)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
        try:
            aba_nome = "Controle de Projeto"
            if aba_nome not in wb.sheetnames:
                return None
            ws = wb[aba_nome]
            resultado: dict[str, Any] = {}
            for campo, (_, celula) in CAMPOS_PROJETO.items():
                try:
                    resultado[campo] = ws[celula].value
                except Exception:
                    resultado[campo] = None
            return resultado
        finally:
            wb.close()
    except Exception:
        return None


def _status_cancelado(status: Any) -> bool:
    return isinstance(status, str) and status.strip().upper() == "CANCELADO"


def listar_projetos(pasta: str) -> list[dict]:
    """
    Varre a pasta, lê metadados de cada .xlsm/.xlsx e retorna lista
    ordenada por DEP, excluindo projetos CANCELADOS.
    """
    if not os.path.isdir(pasta):
        raise FileNotFoundError(f"Pasta não encontrada: {pasta}")

    arquivos = [
        entry.path
        for entry in os.scandir(pasta)
        if entry.is_file()
        and Path(entry.name).suffix.lower() in EXTENSOES_VALIDAS
        and not entry.name.startswith("~$")
    ]

    projetos: list[dict] = []
    for caminho in arquivos:
        dados = _ler_projeto_xml(caminho) or _ler_projeto_openpyxl(caminho)
        if not dados:
            continue
        if _status_cancelado(dados.get("status")):
            continue

        # Normaliza strings
        for k, v in dados.items():
            if isinstance(v, str):
                dados[k] = v.strip()

        dados["arquivo"] = os.path.basename(caminho)
        dados["caminho"] = caminho
        projetos.append(dados)

    # Ordena por DEP (ex: "0002/26")
    projetos.sort(key=lambda p: str(p.get("dep") or ""))
    return projetos
